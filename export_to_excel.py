"""
export_to_excel.py
------------------
Export all Sepsis Atlas tables to a formatted Excel workbook.

Usage
-----
    python export_to_excel.py                    # uses sample data (no DB needed)
    python export_to_excel.py --live             # pulls live data from PostgreSQL
    python export_to_excel.py --live --out my.xlsx
"""

import argparse
import json
import sys
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Sample / fallback data ────────────────────────────────────────────────────
SAMPLE_PAPERS = [
    {
        "id": 1,
        "paper_id": "psofa-prism-2021",
        "title": "Diagnostic accuracy of PRISM III and p-SOFA for mortality prediction in PICU",
        "ingested_at": "2026-05-08 10:00:00",
    }
]

SAMPLE_FINDINGS = [
    {
        "id": 1,
        "paper_id": "psofa-prism-2021",
        "population_type": "Critically ill children (1 month–15 years) admitted to PICU for >24 hours",
        "predictors": [
            "p-SOFA score (maximum within 24 hours; cutoff >2)",
            "PRISM III 24 score (within 24 hours; cutoff >8)",
        ],
        "affected_or_not": True,
        "created_at": "2026-05-08 10:00:05",
    }
]

SAMPLE_EVIDENCE = [
    {
        "id": 1,
        "finding_id": 1,
        "paper_id": "psofa-prism-2021",
        "population_type": "Critically ill children (1 month–15 years) admitted to PICU for >24 hours",
        "predictors": [
            "maximum p-SOFA score (within 24 hours; cutoff >2)",
            "PRISM III 24 score (within 24 hours; cutoff >8)",
        ],
        "affected_or_not": True,
        "sample_size": "286",
        "outcome": "Mortality within 30 days of PICU admission",
        "timing": "Predictors within 24 hours of admission; outcome at 30 days",
        "method": (
            "Single-center cross-validation study (non-probability consecutive sampling). "
            "Discrimination via ROC AUC; diagnostic accuracy from 2×2 tables; p≤0.05 significant."
        ),
        "effect_size": (
            "p-SOFA: AUC=0.81 (95% CI 0.76–0.86), p=0.001. "
            "PRISM III 24: AUC=0.75 (95% CI 0.69–0.81), p=0.001."
        ),
        "performance": (
            "p-SOFA>2: sensitivity=93.87%, specificity=38.21%, accuracy=69.93%. "
            "PRISM III 24>8: sensitivity=55.83%, specificity=77.24%, accuracy=65.03%."
        ),
        "notes": (
            "Median age 24 months (range 1–144). 30-day mortality 57%. "
            "Exclusions: congenital deformity, CPR before admission, death within 12 h."
        ),
        "created_at": "2026-05-08 10:00:10",
    }
]


# ── Live DB fetch ─────────────────────────────────────────────────────────────
def fetch_live_data():
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
    from db.connection import get_conn, put_conn

    conn = get_conn()
    try:
        def query(sql):
            with conn.cursor() as cur:
                cur.execute(sql)
                cols = [d[0] for d in cur.description]
                return [dict(zip(cols, row)) for row in cur.fetchall()]

        papers   = query("SELECT id, paper_id, title, ingested_at FROM papers ORDER BY ingested_at")
        findings = query("SELECT id, paper_id, population_type, predictors, affected_or_not, created_at FROM findings ORDER BY id")
        evidence = query(
            "SELECT id, finding_id, paper_id, population_type, predictors, affected_or_not, "
            "sample_size, outcome, timing, method, effect_size, performance, notes, created_at "
            "FROM evidence ORDER BY id"
        )
        return papers, findings, evidence
    finally:
        put_conn(conn)


# ── Style constants ───────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1F3864")
ALT_FILL  = PatternFill("solid", fgColor="EEF2F7")
YES_FILL  = PatternFill("solid", fgColor="C6EFCE")
NO_FILL   = PatternFill("solid", fgColor="FFC7CE")
HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
WRAP      = Alignment(wrap_text=True, vertical="top")
CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def _hdr(ws, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font, cell.fill, cell.alignment, cell.border = HDR_FONT, HDR_FILL, CENTER, THIN_BORDER
    ws.row_dimensions[1].height = 28


def _cell(ws, r, c, v, is_bool=False, bool_val=None):
    cell = ws.cell(r, c, v)
    cell.font   = BODY_FONT
    cell.border = THIN_BORDER
    cell.alignment = WRAP
    if is_bool:
        cell.fill = YES_FILL if bool_val else NO_FILL
    elif r % 2 == 0:
        cell.fill = ALT_FILL


def _widths(ws, widths):
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def _fmt(v):
    """Stringify dates and lists for Excel cells."""
    if isinstance(v, list):
        return "; ".join(str(x) for x in v)
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    return v


# ── Sheet builders ────────────────────────────────────────────────────────────
def build_overview(wb, papers, findings, evidence):
    ws = wb.active
    ws.title = "Overview"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 50

    ws["A1"].value = "Sepsis Atlas — Database Export"
    ws["A1"].font  = Font(name="Arial", bold=True, size=16, color="1F3864")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 36

    rows = [
        ("Export date",    datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Papers",         len(papers)),
        ("Findings",       len(findings)),
        ("Evidence rows",  len(evidence)),
    ]
    for i, (label, val) in enumerate(rows, 3):
        ws.cell(i, 1).value = label
        ws.cell(i, 1).font  = Font(name="Arial", bold=True, size=11)
        ws.cell(i, 2).value = val
        ws.cell(i, 2).font  = Font(name="Arial", size=11)

    ws["A8"].value = "Sheets"
    ws["A8"].font  = Font(name="Arial", bold=True, size=11, color="1F3864")
    desc = [
        ("Papers",   "One row per ingested paper"),
        ("Findings", "High-level extraction findings per paper"),
        ("Evidence", "Detailed evidence records with effect sizes, performance, methods"),
    ]
    for i, (name, d) in enumerate(desc, 9):
        ws.cell(i, 1).value = name
        ws.cell(i, 1).font  = Font(name="Arial", bold=True, size=10)
        ws.cell(i, 2).value = d
        ws.cell(i, 2).font  = Font(name="Arial", size=10)


def build_papers(wb, papers):
    ws = wb.create_sheet("Papers")
    headers = ["ID", "Paper ID", "Title", "Ingested At"]
    _hdr(ws, headers)
    _widths(ws, [6, 28, 65, 22])
    for r, p in enumerate(papers, 2):
        for c, v in enumerate([p["id"], p["paper_id"], p["title"], _fmt(p["ingested_at"])], 1):
            _cell(ws, r, c, v)
        ws.row_dimensions[r].height = 18


def build_findings(wb, findings):
    ws = wb.create_sheet("Findings")
    headers = ["ID", "Paper ID", "Population Type", "Predictors", "Associated", "Created At"]
    _hdr(ws, headers)
    _widths(ws, [6, 28, 45, 58, 14, 22])
    for r, f in enumerate(findings, 2):
        af = f["affected_or_not"]
        vals = [f["id"], f["paper_id"], f["population_type"], _fmt(f["predictors"]),
                "Yes" if af else "No", _fmt(f["created_at"])]
        for c, v in enumerate(vals, 1):
            _cell(ws, r, c, v, is_bool=(c == 5), bool_val=af)
        ws.row_dimensions[r].height = 45


def build_evidence(wb, evidence):
    ws = wb.create_sheet("Evidence")
    headers = [
        "ID", "Finding ID", "Paper ID", "Population Type", "Predictors",
        "Associated", "Sample Size", "Outcome", "Timing", "Method",
        "Effect Size", "Performance", "Notes", "Created At",
    ]
    _hdr(ws, headers)
    _widths(ws, [6, 10, 24, 38, 48, 12, 12, 38, 38, 52, 48, 58, 58, 22])
    for r, e in enumerate(evidence, 2):
        af = e["affected_or_not"]
        vals = [
            e["id"], e["finding_id"], e["paper_id"], e["population_type"],
            _fmt(e["predictors"]), "Yes" if af else "No",
            e["sample_size"], e["outcome"], e["timing"], e["method"],
            e["effect_size"], e["performance"], e["notes"], _fmt(e["created_at"]),
        ]
        for c, v in enumerate(vals, 1):
            _cell(ws, r, c, v, is_bool=(c == 6), bool_val=af)
        ws.row_dimensions[r].height = 85


# ── Entry point ───────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--live", action="store_true", help="Pull data from PostgreSQL")
    parser.add_argument("--out", default="sepsis_atlas_export.xlsx", help="Output filename")
    args = parser.parse_args()

    if args.live:
        print("[export] Fetching live data from PostgreSQL…")
        papers, findings, evidence = fetch_live_data()
        print(f"[export] {len(papers)} papers, {len(findings)} findings, {len(evidence)} evidence rows.")
    else:
        print("[export] Using sample data (run with --live to export from DB).")
        papers, findings, evidence = SAMPLE_PAPERS, SAMPLE_FINDINGS, SAMPLE_EVIDENCE

    wb = Workbook()
    build_overview(wb, papers, findings, evidence)
    build_papers(wb, papers)
    build_findings(wb, findings)
    build_evidence(wb, evidence)

    wb.save(args.out)
    print(f"[export] Saved → {args.out}")


if __name__ == "__main__":
    main()
